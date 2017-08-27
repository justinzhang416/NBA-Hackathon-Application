from openpyxl import load_workbook
from openpyxl import Workbook
import operator
import copy

"""A Team class to represent a team's name, record, and record vs other teams"""
class Team:
    def __init__(self, name):
        self.name = name
        self.w = 0
        self.l = 0
        self.date = ""
        self.vsRecord = {}
    def __lt__(self, other):
        if self.l == other.l:
            return self.vsRecord[other.name].l < other.vsRecord[self.name].l
        return self.l < other.l
    
"""A Record class to represent wins and losses"""
class Record:
    def __init__(self):
        self.w = 0
        self.l = 0
        
'''Helper function that calculates standings and determines which teams
have been eliminated'''
def eliminated(date):
    eastStanding.sort()
    westStanding.sort()

    #If eighths seeds current wins are greater than if current team wins
    #the rest of their games, current team is automatically eliminated
    for i in range(8, 15):
        eighthseed=eastStanding[7]
        team = eastStanding[i]
        if eighthseed.w>(team.w+(82-team.l-team.w)) and team.date=="":
            team.date=date
        if team.w + team.l == 82 and team.date =="":
            team.date=date
            
        eighthseed=westStanding[7]
        team = westStanding[i]
        if eighthseed.w>(team.w+(82-team.l-team.w)) and team.date=="":
            team.date=date
        if team.w + team.l == 82 and team.date =="":
            team.date=date

#Import data
data = load_workbook('Analytics_Attachment.xlsx',read_only=True)
teamData = data['Division_Info']
ws = data['2016_17_NBA_Scores']

east = {}
west = {}
total = {}
VS = {}

#Create dictionary of teams
for i in range(2, 32):
    name = teamData.cell(row=i,column=1).value
    conf = teamData.cell(row=i,column=3).value
    currentTeam = Team(name)
    total[name] = currentTeam
    VS[name] = Record()
    if(conf == "East"):
        east[name] = currentTeam
    else:
        west[name] = currentTeam
        
#Assign each team a blank record vs all other teams       
for key in total:
    total[key].vsRecord = copy.deepcopy(VS)

#Initialize date, and standings. Both are currently meaningless.
currentDate = ""
eastStanding = list(east.values())
westStanding = list(west.values())

for i in range(2, 1232):
    date = ws.cell(row=i,column=1).value
    home = ws.cell(row=i,column=2).value
    away = ws.cell(row=i,column=3).value
    winner = ws.cell(row=i,column=6).value

    #Once new date detected, calculate new standings and check for
    #eliminated teams
    if(currentDate != date):
        eliminated(currentDate)
        
    currentDate = date       
    if(winner == "Home"):
        winTeam = total[home]
        winTeam.w += 1
        winTeam.vsRecord[away].w += 1
        loseTeam = total[away]
        loseTeam.l += 1
        loseTeam.vsRecord[home].l += 1
    else:
        winTeam = total[away]
        winTeam.w += 1
        winTeam.vsRecord[away].w += 1
        loseTeam = total[home]
        loseTeam.l += 1
        loseTeam.vsRecord[home].l += 1
#Call eliminated on final date
eliminated(currentDate)

#Write results to file
wb = Workbook()
result = wb.create_sheet(title="Results")
result.cell(row=1, column=1).value = "Team"
result.cell(row=1, column=2).value = "Date Eliminated"


print("Eastern Conference")
for item in eastStanding:
    result.cell(row=eastStanding.index(item)+2, column=1).value = item.name
    result.cell(row=eastStanding.index(item)+2, column=2).value = item.date
    print(str(eastStanding.index(item)+1) + ": " + item.name+ ".  W_L Record is: "
          + str(item.w) + "-" + str(item.l) +" Date eliminated is " + str(item.date))

print ("\nWestern Conference")
for item in westStanding:
    result.cell(row=westStanding.index(item)+17, column=1).value = item.name
    result.cell(row=westStanding.index(item)+17, column=2).value = item.date
    print( str(westStanding.index(item)+1) + ": " + item.name+ ". W_L Record is: "
          + str(item.w) + "-" + str(item.l) +" Date eliminated is " + str(item.date))
wb.save('EliminationResults.xlsx')

